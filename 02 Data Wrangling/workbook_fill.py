# This script will copy data from each year of observations into the master workbook for a
# given gauge site.  The yearly data consists of five columns, each with a maximum of 8785
# rows.  The first page of the master workbook is the summary page; data needs to be copied
# into the appropriately-named sheet following the first page.  The template for the master
# workbook creates a sheet for each year between 1995 and 2015 (inclusive).  Each master
# workbook is named according to the numeric code of its gauge.
#
# The biggest problem with this script is that it loses the two tables that were part of the
# original master workbook template.  Not sure how to keep that from happening.


from openpyxl import Workbook
from openpyxl import load_workbook
import os
import glob
import csv

data_folder = "/Users/kate/Box Sync/PuertoRicoSedimentYields_SharedFolder/Kate's Work/Data/"



def main():
    siteno = raw_input("Input USGS site number: ")
    gauge_folder = "{0}{1}/".format(data_folder, siteno)
    masterWorkbook = "{0}/{1}.xlsx".format(gauge_folder, siteno)

    # Proceed if there's a data folder that matches that site number
    if os.path.isdir(gauge_folder):

        # This checks to see if a master workbook already exists in the gauge data folder.
        # If not, it makes a new one using the inst_dis template.
        if not os.path.isfile(masterWorkbook):
            print("Creating new master workbook for site {}".format(siteno))
            wb = load_workbook(data_folder + 'inst_dis.xltx')
            wb.template = False
            wb.save(masterWorkbook)
            print("New master workbook created.")

        # Make the master workbook the active workbook.
        wb = load_workbook(masterWorkbook)

        # This is pretty accurate at guessing the data type per cell.
        wb.guess_types = True

        for csvfile in glob.glob(os.path.join(gauge_folder, '*.csv')):
            # This gets the year associated with the file in question
            year = csvfile[-8:-4]

            # Use this to set the active sheet in the master workbook
            ws = wb[year]

            with open(csvfile, 'rb') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        if c < 5: # We only want the first five columns
                            # Write the csv file content into appropriate sheet.
                            ws.cell(row=r+1, column=c+1, value=col)

            print("{} Complete".format(year))

        print("Saving data, please be patient.")
        wb.save(masterWorkbook)
        print("Data saved to master workbook.  Process complete.")

    else:
        print("No data found for that site number.")


main()

